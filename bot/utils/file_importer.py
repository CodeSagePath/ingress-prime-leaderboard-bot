#!/usr/bin/env python3
"""
Resilient File Importer Module for Ingress Prime Leaderboard Bot
Supports multiple file formats with automatic detection and robust error handling
"""

import csv
import json
import logging
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
import io
import chardet

from .data_validator import DataValidator, validate_players_data

logger = logging.getLogger(__name__)


class FileImportError(Exception):
    """Custom exception for file import errors"""
    pass


class FileImporter:
    """
    Handle various file formats for player data with automatic format detection
    and comprehensive error handling
    """

    SUPPORTED_FORMATS = {
        '.json': 'json',
        '.csv': 'csv',
        '.xlsx': 'excel',
        '.xls': 'excel',
        '.txt': 'text',
        '.tsv': 'tsv'
    }

    def __init__(self, max_file_size_mb: int = 10):
        """
        Initialize file importer

        Args:
            max_file_size_mb: Maximum file size in megabytes
        """
        self.max_file_size_mb = max_file_size_mb
        self.max_file_size_bytes = max_file_size_mb * 1024 * 1024

    @staticmethod
    def detect_encoding(file_path: Union[str, Path]) -> str:
        """
        Detect file encoding automatically

        Args:
            file_path: Path to the file

        Returns:
            Detected encoding string
        """
        try:
            with open(file_path, 'rb') as f:
                raw_data = f.read(10000)  # Read first 10KB for detection
                result = chardet.detect(raw_data)
                encoding = result.get('encoding', 'utf-8')
                confidence = result.get('confidence', 0)

                logger.info(f"Detected encoding: {encoding} (confidence: {confidence:.2f})")

                # Fallback to utf-8 if confidence is low
                if confidence < 0.7:
                    logger.warning(f"Low encoding confidence ({confidence:.2f}), trying utf-8")
                    return 'utf-8'

                return encoding
        except Exception as e:
            logger.warning(f"Encoding detection failed: {e}, defaulting to utf-8")
            return 'utf-8'

    def validate_file(self, file_path: Union[str, Path]) -> bool:
        """
        Validate file before processing

        Args:
            file_path: Path to the file

        Returns:
            True if file is valid for processing

        Raises:
            FileImportError: If file is invalid
        """
        file_path = Path(file_path)

        # Check if file exists
        if not file_path.exists():
            raise FileImportError(f"File does not exist: {file_path}")

        # Check file size
        file_size = file_path.stat().st_size
        if file_size > self.max_file_size_bytes:
            raise FileImportError(
                f"File too large: {file_size / (1024*1024):.1f}MB "
                f"(max allowed: {self.max_file_size_mb}MB)"
            )

        # Check if file is empty
        if file_size == 0:
            raise FileImportError("File is empty")

        # Check file format
        file_ext = file_path.suffix.lower()
        if file_ext not in self.SUPPORTED_FORMATS:
            supported_formats = ', '.join(self.SUPPORTED_FORMATS.keys())
            raise FileImportError(
                f"Unsupported file format: {file_ext}. "
                f"Supported formats: {supported_formats}"
            )

        return True

    @staticmethod
    def detect_format(file_path: Union[str, Path]) -> Optional[str]:
        """
        Detect file format from extension and content

        Args:
            file_path: Path to the file

        Returns:
            Detected format string
        """
        file_path = Path(file_path)
        ext = file_path.suffix.lower()

        # Basic extension mapping
        format_map = {
            '.json': 'json',
            '.csv': 'csv',
            '.xlsx': 'excel',
            '.xls': 'excel',
            '.txt': 'text',
            '.tsv': 'tsv'
        }

        detected_format = format_map.get(ext)

        # Additional content-based detection for ambiguous files
        if not detected_format and ext in ['.txt', '.dat']:
            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    first_line = f.readline().strip()

                # Check if it looks like CSV/TSV
                if ',' in first_line or '\t' in first_line:
                    detected_format = 'tsv' if '\t' in first_line else 'csv'
                # Check if it looks like JSON
                elif first_line.startswith('{') or first_line.startswith('['):
                    detected_format = 'json'
                else:
                    detected_format = 'text'

                logger.info(f"Content-based detection: {ext} file is {detected_format}")

            except Exception as e:
                logger.warning(f"Content detection failed: {e}")
                detected_format = 'text'

        return detected_format

    def import_json(self, file_path: Union[str, Path]) -> List[Dict[str, Any]]:
        """
        Import JSON file with comprehensive error handling

        Args:
            file_path: Path to JSON file

        Returns:
            List of dictionaries containing player data

        Raises:
            FileImportError: If JSON parsing fails
        """
        try:
            encoding = self.detect_encoding(file_path)

            with open(file_path, 'r', encoding=encoding) as f:
                data = json.load(f)

            # Handle different JSON structures
            if isinstance(data, dict):
                # Single record
                return [data]
            elif isinstance(data, list):
                # Multiple records
                if not data:
                    logger.warning("JSON file contains empty array")
                return data
            else:
                raise FileImportError("JSON file must contain a dictionary or array of dictionaries")

        except json.JSONDecodeError as e:
            raise FileImportError(f"Invalid JSON format: {str(e)}")
        except Exception as e:
            raise FileImportError(f"Error reading JSON file: {str(e)}")

    def import_csv(self, file_path: Union[str, Path]) -> List[Dict[str, Any]]:
        """
        Import CSV file with flexible column mapping and encoding detection

        Args:
            file_path: Path to CSV file

        Returns:
            List of dictionaries containing player data

        Raises:
            FileImportError: If CSV parsing fails
        """
        try:
            encoding = self.detect_encoding(file_path)

            # First, detect delimiter
            with open(file_path, 'r', encoding=encoding, newline='') as f:
                sample = f.read(1024)
                sniffer = csv.Sniffer()

                try:
                    delimiter = sniffer.sniff(sample).delimiter
                except csv.Error:
                    # Fallback delimiters
                    delimiter = '\t' if '\t' in sample else (',' if ',' in sample else ';')

            logger.info(f"Detected CSV delimiter: '{delimiter}'")

            # Read CSV with proper parameters
            with open(file_path, 'r', encoding=encoding, newline='') as f:
                # Handle files with BOM
                if encoding.lower() == 'utf-8' and sample.startswith('\ufeff'):
                    f.seek(0)
                    content = f.read().lstrip('\ufeff')
                    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
                else:
                    f.seek(0)
                    reader = csv.DictReader(f, delimiter=delimiter)

                if not reader.fieldnames:
                    raise FileImportError("CSV file has no headers")

                logger.info(f"CSV headers detected: {reader.fieldnames}")

                # Process rows
                data = []
                empty_rows = 0
                for row_num, row in enumerate(reader, 1):
                    # Skip empty rows
                    if all(not value.strip() for value in row.values()):
                        empty_rows += 1
                        continue

                    # Clean and normalize row data
                    cleaned_row = {}
                    for key, value in row.items():
                        if key:  # Skip empty column names
                            cleaned_key = key.strip()
                            cleaned_value = value.strip() if isinstance(value, str) else value
                            cleaned_row[cleaned_key] = cleaned_value

                    if cleaned_row:
                        data.append(cleaned_row)

                if empty_rows > 0:
                    logger.warning(f"Skipped {empty_rows} empty rows in CSV file")

                return data

        except csv.Error as e:
            raise FileImportError(f"CSV parsing error: {str(e)}")
        except UnicodeDecodeError as e:
            raise FileImportError(f"Encoding error: {str(e)}. Try saving the file as UTF-8.")
        except Exception as e:
            raise FileImportError(f"Error reading CSV file: {str(e)}")

    def import_excel(self, file_path: Union[str, Path]) -> List[Dict[str, Any]]:
        """
        Import Excel file with error handling and multiple sheet support

        Args:
            file_path: Path to Excel file

        Returns:
            List of dictionaries containing player data

        Raises:
            FileImportError: If Excel parsing fails
        """
        try:
            # Try using pandas first (more robust)
            try:
                # Read all sheets
                excel_file = pd.ExcelFile(file_path)

                # Find the sheet with the most data
                best_sheet = None
                max_rows = 0

                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    row_count = len(df.dropna(how='all'))  # Count non-empty rows

                    if row_count > max_rows:
                        max_rows = row_count
                        best_sheet = sheet_name

                logger.info(f"Using Excel sheet '{best_sheet}' with {max_rows} data rows")

                # Read the best sheet
                df = pd.read_excel(excel_file, sheet_name=best_sheet)

                # Drop completely empty rows
                df = df.dropna(how='all')

                # Drop completely empty columns
                df = df.dropna(axis=1, how='all')

                # Convert to list of dictionaries
                data = df.to_dict('records')

                # Clean the data
                cleaned_data = []
                for row in data:
                    cleaned_row = {}
                    for key, value in row.items():
                        if pd.notna(key) and pd.notna(value):
                            # Handle pandas data types
                            if pd.isna(value):
                                continue
                            elif isinstance(value, (pd.Timestamp, pd.Timedelta)):
                                cleaned_row[str(key).strip()] = str(value)
                            else:
                                cleaned_row[str(key).strip()] = value
                    if cleaned_row:
                        cleaned_data.append(cleaned_row)

                return cleaned_data

            except ImportError:
                # Fallback to openpyxl if pandas is not available
                logger.warning("pandas not available, falling back to openpyxl")
                from openpyxl import load_workbook

                workbook = load_workbook(file_path, read_only=True)
                sheet_names = workbook.sheetnames

                # Find the sheet with the most data
                best_sheet_name = sheet_names[0]
                max_rows = 0

                for sheet_name in sheet_names:
                    sheet = workbook[sheet_name]
                    # Count non-empty rows in first column
                    row_count = sum(1 for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True) if any(cell is not None for cell in row))
                    if row_count > max_rows:
                        max_rows = row_count
                        best_sheet_name = sheet_name

                logger.info(f"Using Excel sheet '{best_sheet_name}' with {max_rows} data rows")

                sheet = workbook[best_sheet_name]

                # Get headers from first row
                headers = [cell.value for cell in sheet[1] if cell.value is not None]
                if not headers:
                    raise FileImportError("Excel file has no headers in first row")

                logger.info(f"Excel headers: {headers}")

                # Read data rows
                data = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None for cell in row):  # Skip empty rows
                        record = dict(zip(headers, row))
                        # Clean the record
                        cleaned_record = {
                            str(key).strip(): value for key, value in record.items()
                            if key is not None and value is not None
                        }
                        if cleaned_record:
                            data.append(cleaned_record)

                workbook.close()
                return data

        except ImportError as e:
            raise FileImportError(
                f"Neither pandas nor openpyxl is available for Excel support. "
                f"Please install with: pip install pandas openpyxl. Original error: {str(e)}"
            )
        except Exception as e:
            raise FileImportError(f"Error reading Excel file: {str(e)}")

    def import_text(self, file_path: Union[str, Path]) -> List[Dict[str, Any]]:
        """
        Import text file with various delimiters

        Args:
            file_path: Path to text file

        Returns:
            List of dictionaries containing player data

        Raises:
            FileImportError: If text parsing fails
        """
        try:
            encoding = self.detect_encoding(file_path)

            with open(file_path, 'r', encoding=encoding) as f:
                content = f.read()

            # Try to detect delimiter and structure
            lines = content.strip().split('\n')

            if not lines:
                raise FileImportError("Text file is empty")

            # Skip empty lines
            lines = [line.strip() for line in lines if line.strip()]

            # Try to parse as delimiter-separated values
            first_line = lines[0]

            # Detect delimiter
            if '\t' in first_line:
                delimiter = '\t'
                format_type = 'tsv'
            elif '|' in first_line:
                delimiter = '|'
                format_type = 'pipe'
            elif ';' in first_line and ',' not in first_line:
                delimiter = ';'
                format_type = 'semicolon'
            else:
                delimiter = ','
                format_type = 'csv'

            logger.info(f"Detected text format: {format_type} with delimiter '{delimiter}'")

            # Parse using the detected delimiter
            if len(lines) == 1:
                # Single line - treat as key-value pairs
                return [dict(item.split(':', 1) if ':' in item else item.split(delimiter, 1) for item in first_line.split(delimiter))]
            else:
                # Multiple lines - treat as CSV-like
                import io
                reader = csv.DictReader(io.StringIO('\n'.join(lines)), delimiter=delimiter)

                data = []
                for row in reader:
                    cleaned_row = {k.strip(): v.strip() for k, v in row.items() if k and v}
                    if cleaned_row:
                        data.append(cleaned_row)

                return data

        except Exception as e:
            raise FileImportError(f"Error reading text file: {str(e)}")

    def import_file(self, file_path: Union[str, Path], validate: bool = True, strict_validation: bool = False) -> Dict[str, Any]:
        """
        Auto-detect and import file with comprehensive validation

        Args:
            file_path: Path to file
            validate: Whether to validate the data after import
            strict_validation: Use strict validation mode

        Returns:
            Dictionary with import results including validation status

        Raises:
            FileImportError: If file import fails
        """
        file_path = Path(file_path)

        # Validate file
        self.validate_file(file_path)

        # Detect format
        file_format = self.detect_format(file_path)
        if not file_format:
            raise FileImportError(f"Could not detect file format for: {file_path}")

        logger.info(f"Importing {file_format} file: {file_path}")

        try:
            # Import based on format
            if file_format == 'json':
                data = self.import_json(file_path)
            elif file_format == 'csv':
                data = self.import_csv(file_path)
            elif file_format == 'excel':
                data = self.import_excel(file_path)
            elif file_format in ['text', 'tsv']:
                data = self.import_text(file_path)
            else:
                raise FileImportError(f"No handler for format: {file_format}")

            logger.info(f"Successfully imported {len(data)} records from {file_path}")

            # Validate data if requested
            validation_result = None
            if validate and data:
                try:
                    validation_result = validate_players_data(data, strict_validation)
                except Exception as e:
                    logger.error(f"Validation failed: {e}")
                    validation_result = {
                        'valid': False,
                        'total_count': len(data),
                        'valid_count': 0,
                        'invalid_count': len(data),
                        'summary_errors': [f"Validation failed: {str(e)}"],
                        'global_warnings': []
                    }

            return {
                'success': True,
                'file_path': str(file_path),
                'file_format': file_format,
                'total_records': len(data),
                'data': data,
                'validation': validation_result,
                'errors': [],
                'warnings': []
            }

        except Exception as e:
            logger.error(f"File import failed: {e}")
            return {
                'success': False,
                'file_path': str(file_path),
                'file_format': file_format,
                'total_records': 0,
                'data': [],
                'validation': None,
                'errors': [str(e)],
                'warnings': []
            }

    @staticmethod
    def get_supported_formats() -> str:
        """Get a formatted string of supported file formats"""
        formats = []
        for ext, format_name in FileImporter.SUPPORTED_FORMATS.items():
            formats.append(f"{ext} ({format_name})")
        return ", ".join(formats)


# Convenience function for quick file import
def import_player_data_file(file_path: Union[str, Path], validate: bool = True, strict_validation: bool = False) -> Dict[str, Any]:
    """
    Quick file import function

    Args:
        file_path: Path to the file
        validate: Whether to validate imported data
        strict_validation: Use strict validation mode

    Returns:
        Dictionary with import results
    """
    importer = FileImporter()
    return importer.import_file(file_path, validate, strict_validation)